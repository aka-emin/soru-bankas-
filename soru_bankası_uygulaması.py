import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QTextEdit, QRadioButton, QButtonGroup,
    QPushButton, QListWidget, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QTabWidget, QGroupBox
)
from PyQt5.QtCore import Qt
from openpyxl import Workbook, load_workbook
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtGui import QTextDocument, QFont

class ExamSystemApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Öğrenci Sınav Sistemi v2.0")
        self.setGeometry(100, 100, 900, 600)
        
        # Ana widget ve layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        
        # Sekmeler oluştur
        self.create_tabs()
        
        # Soru verileri
        self.questions = []
        
        # Stilleri ayarla
        self.set_styles()
    
    def set_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTabWidget::pane {
                border: 1px solid #ccc;
                background: white;
            }
            QTabBar::tab {
                background: #e0e0e0;
                padding: 8px;
                border: 1px solid #ccc;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background: white;
                margin-bottom: -1px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton#export_btn {
                background-color: #2196F3;
            }
            QPushButton#export_btn:hover {
                background-color: #0b7dda;
            }
            QPushButton#print_btn {
                background-color: #ff9800;
            }
            QPushButton#print_btn:hover {
                background-color: #e68a00;
            }
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
    
    def create_tabs(self):
        self.tabs = QTabWidget()
        self.main_layout.addWidget(self.tabs)
        
        # Soru Ekleme Sekmesi
        self.create_question_tab()
        
        # Sınav Oluşturma Sekmesi
        self.create_exam_tab()
        
        # İstatistikler Sekmesi
        self.create_stats_tab()
    
    def create_question_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Soru Giriş Grubu
        question_group = QGroupBox("Yeni Soru Ekle")
        question_layout = QVBoxLayout(question_group)
        
        # Soru metni
        self.question_text = QTextEdit()
        self.question_text.setPlaceholderText("Soruyu buraya yazın...")
        question_layout.addWidget(QLabel("Soru Metni:"))
        question_layout.addWidget(self.question_text)
        
        # Seçenekler
        options_group = QGroupBox("Seçenekler")
        options_layout = QVBoxLayout(options_group)
        
        self.option_a = QLineEdit()
        self.option_b = QLineEdit()
        self.option_c = QLineEdit()
        self.option_d = QLineEdit()
        self.option_e = QLineEdit()
        
        options_layout.addWidget(QLabel("A Seçeneği:"))
        options_layout.addWidget(self.option_a)
        options_layout.addWidget(QLabel("B Seçeneği:"))
        options_layout.addWidget(self.option_b)
        options_layout.addWidget(QLabel("C Seçeneği:"))
        options_layout.addWidget(self.option_c)
        options_layout.addWidget(QLabel("D Seçeneği:"))
        options_layout.addWidget(self.option_d)
        options_layout.addWidget(QLabel("E Seçeneği:"))
        options_layout.addWidget(self.option_e)
        
        # Doğru cevap
        self.correct_answer = QButtonGroup()
        self.radio_a = QRadioButton("A")
        self.radio_b = QRadioButton("B")
        self.radio_c = QRadioButton("C")
        self.radio_d = QRadioButton("D")
        self.radio_e = QRadioButton("E")
        
        self.correct_answer.addButton(self.radio_a)
        self.correct_answer.addButton(self.radio_b)
        self.correct_answer.addButton(self.radio_c)
        self.correct_answer.addButton(self.radio_d)
        self.correct_answer.addButton(self.radio_e)
        
        answer_layout = QHBoxLayout()
        answer_layout.addWidget(QLabel("Doğru Cevap:"))
        answer_layout.addWidget(self.radio_a)
        answer_layout.addWidget(self.radio_b)
        answer_layout.addWidget(self.radio_c)
        answer_layout.addWidget(self.radio_d)
        answer_layout.addWidget(self.radio_e)
        answer_layout.addStretch()
        
        # Butonlar
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Soruyu Ekle")
        add_btn.clicked.connect(self.add_question)
        clear_btn = QPushButton("Temizle")
        clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(clear_btn)
        
        # Tüm widget'ları yerleştir
        question_layout.addWidget(options_group)
        question_layout.addLayout(answer_layout)
        question_layout.addLayout(btn_layout)
        
        # Soru Listesi
        self.question_list = QListWidget()
        self.question_list.setAlternatingRowColors(True)
        
        # İşlem Butonları
        action_layout = QHBoxLayout()
        export_btn = QPushButton("Excel'e Aktar")
        export_btn.setObjectName("export_btn")
        export_btn.clicked.connect(self.export_to_excel)
        import_btn = QPushButton("Excel'den İçe Aktar")
        import_btn.setObjectName("export_btn")
        import_btn.clicked.connect(self.import_from_excel)
        action_layout.addWidget(export_btn)
        action_layout.addWidget(import_btn)
        
        layout.addWidget(question_group)
        layout.addWidget(QLabel("Eklenen Sorular:"))
        layout.addWidget(self.question_list)
        layout.addLayout(action_layout)
        
        self.tabs.addTab(tab, "Soru Ekleme")
    
    def create_exam_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Kontrol Butonları
        control_layout = QHBoxLayout()
        load_btn = QPushButton("Soru Yükle")
        load_btn.clicked.connect(self.load_questions)
        print_btn = QPushButton("PDF Oluştur")
        print_btn.setObjectName("print_btn")
        print_btn.clicked.connect(self.generate_pdf)
        control_layout.addWidget(load_btn)
        control_layout.addWidget(print_btn)
        
        # Soru Tablosu
        self.question_table = QTableWidget()
        self.question_table.setColumnCount(7)
        self.question_table.setHorizontalHeaderLabels(["Soru", "A", "B", "C", "D", "E", "Doğru Cevap"])
        self.question_table.horizontalHeader().setStretchLastSection(True)
        
        # Sınav Ayarları
        settings_group = QGroupBox("Sınav Ayarları")
        settings_layout = QHBoxLayout(settings_group)
        
        self.exam_title = QLineEdit()
        self.exam_title.setPlaceholderText("Sınav Başlığı")
        self.exam_date = QLineEdit()
        self.exam_date.setPlaceholderText("Sınav Tarihi")
        
        settings_layout.addWidget(QLabel("Başlık:"))
        settings_layout.addWidget(self.exam_title)
        settings_layout.addWidget(QLabel("Tarih:"))
        settings_layout.addWidget(self.exam_date)
        
        layout.addLayout(control_layout)
        layout.addWidget(settings_group)
        layout.addWidget(self.question_table)
        
        self.tabs.addTab(tab, "Sınav Oluştur")
    
    def create_stats_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # İstatistik Bilgileri
        stats_label = QLabel("Sistem İstatistikleri")
        stats_label.setAlignment(Qt.AlignCenter)
        stats_label.setFont(QFont("Arial", 14, QFont.Bold))
        
        stats_info = QLabel(
            "Toplam Soru Sayısı: 0\n"
            "Oluşturulan Sınav Sayısı: 0\n"
            "En Çok Kullanılan Soru Tipi: -"
        )
        stats_info.setAlignment(Qt.AlignLeft)
        
        # Grafik Alanı (simülasyon)
        graph_label = QLabel("Soru Dağılım Grafiği (Simülasyon)")
        graph_label.setAlignment(Qt.AlignCenter)
        graph_label.setStyleSheet("background-color: #e0e0e0; padding: 50px;")
        
        layout.addWidget(stats_label)
        layout.addWidget(stats_info)
        layout.addWidget(graph_label)
        layout.addStretch()
        
        self.tabs.addTab(tab, "İstatistikler")
    
    def add_question(self):
        question = self.question_text.toPlainText().strip()
        options = [
            self.option_a.text().strip(),
            self.option_b.text().strip(),
            self.option_c.text().strip(),
            self.option_d.text().strip(),
            self.option_e.text().strip()
        ]
        
        if not question:
            QMessageBox.warning(self, "Uyarı", "Soru metni boş olamaz!")
            return
        
        if any(not opt for opt in options):
            QMessageBox.warning(self, "Uyarı", "Tüm seçenekler doldurulmalıdır!")
            return
        
        correct_answer = None
        if self.radio_a.isChecked(): correct_answer = "A"
        elif self.radio_b.isChecked(): correct_answer = "B"
        elif self.radio_c.isChecked(): correct_answer = "C"
        elif self.radio_d.isChecked(): correct_answer = "D"
        elif self.radio_e.isChecked(): correct_answer = "E"
        
        if not correct_answer:
            QMessageBox.warning(self, "Uyarı", "Doğru cevap seçilmelidir!")
            return
        
        # Soruyu listeye ekle
        question_data = {
            "text": question,
            "options": options,
            "correct": correct_answer
        }
        self.questions.append(question_data)
        
        # ListWidget'a ekle
        item_text = f"{question[:50]}..." if len(question) > 50 else question
        self.question_list.addItem(f"{len(self.questions)}. {item_text} (Doğru: {correct_answer})")
        
        # Formu temizle
        self.clear_form()
        
        QMessageBox.information(self, "Başarılı", "Soru başarıyla eklendi!")
    
    def clear_form(self):
        self.question_text.clear()
        self.option_a.clear()
        self.option_b.clear()
        self.option_c.clear()
        self.option_d.clear()
        self.option_e.clear()
        self.correct_answer.setExclusive(False)
        for btn in self.correct_answer.buttons():
            btn.setChecked(False)
        self.correct_answer.setExclusive(True)
    
    def export_to_excel(self):
        if not self.questions:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek soru bulunamadı!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel Dosyası Olarak Kaydet", "", "Excel Files (*.xlsx)")
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sorular"
            
            # Başlık satırı
            headers = ["Soru", "A Seçeneği", "B Seçeneği", "C Seçeneği", 
                      "D Seçeneği", "E Seçeneği", "Doğru Cevap"]
            ws.append(headers)
            
            # Soruları ekle
            for q in self.questions:
                row = [q['text']] + q['options'] + [q['correct']]
                ws.append(row)
            
            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", "Sorular Excel dosyasına kaydedildi!")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilirken hata oluştu:\n{str(e)}")
    
    def import_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "", "Excel Files (*.xlsx)")
        
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            self.questions.clear()
            self.question_list.clear()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:
                    question_data = {
                        "text": row[0],
                        "options": list(row[1:6]),
                        "correct": row[6]
                    }
                    self.questions.append(question_data)
                    
                    item_text = f"{row[0][:50]}..." if len(row[0]) > 50 else row[0]
                    self.question_list.addItem(f"{len(self.questions)}. {item_text} (Doğru: {row[6]})")
            
            QMessageBox.information(self, "Başarılı", f"{len(self.questions)} soru başarıyla yüklendi!")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya yüklenirken hata oluştu:\n{str(e)}")
    
    def load_questions(self):
        if not self.questions:
            QMessageBox.warning(self, "Uyarı", "Yüklenecek soru bulunamadı!")
            return
        
        self.question_table.setRowCount(len(self.questions))
        
        for i, q in enumerate(self.questions):
            self.question_table.setItem(i, 0, QTableWidgetItem(q['text']))
            self.question_table.setItem(i, 1, QTableWidgetItem(q['options'][0]))
            self.question_table.setItem(i, 2, QTableWidgetItem(q['options'][1]))
            self.question_table.setItem(i, 3, QTableWidgetItem(q['options'][2]))
            self.question_table.setItem(i, 4, QTableWidgetItem(q['options'][3]))
            self.question_table.setItem(i, 5, QTableWidgetItem(q['options'][4]))
            self.question_table.setItem(i, 6, QTableWidgetItem(q['correct']))
    
    def generate_pdf(self):
        if self.question_table.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "PDF oluşturmak için önce soruları yükleyin!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "PDF Olarak Kaydet", "", "PDF Files (*.pdf)")
        
        if not file_path:
            return
        
        # HTML içeriği oluştur
        html = """
        <html>
        <head>
        <style>
            body { font-family: Arial; }
            h1 { color: #2c3e50; text-align: center; }
            h2 { color: #34495e; border-bottom: 1px solid #eee; padding-bottom: 5px; }
            table { width: 100%; border-collapse: collapse; margin-top: 20px; }
            th { background-color: #3498db; color: white; padding: 8px; text-align: left; }
            td { padding: 8px; border-bottom: 1px solid #ddd; }
            .exam-info { margin-bottom: 20px; }
            .footer { margin-top: 30px; text-align: right; font-style: italic; }
        </style>
        </head>
        <body>
        """
        
        # Sınav bilgileri
        exam_title = self.exam_title.text() or "Genel Sınav"
        exam_date = self.exam_date.text() or "Belirtilmemiş"
        
        html += f"""
        <h1>{exam_title}</h1>
        <div class="exam-info">
            <p><strong>Tarih:</strong> {exam_date}</p>
            <p><strong>Soru Sayısı:</strong> {self.question_table.rowCount()}</p>
        </div>
        
        <h2>Sorular</h2>
        <table>
            <tr>
                <th>#</th>
                <th>Soru</th>
                <th>Seçenekler</th>
                <th>Doğru Cevap</th>
            </tr>
        """
        
        # Soruları ekle
        for row in range(self.question_table.rowCount()):
            question = self.question_table.item(row, 0).text()
            options = [
                f"A) {self.question_table.item(row, 1).text()}",
                f"B) {self.question_table.item(row, 2).text()}",
                f"C) {self.question_table.item(row, 3).text()}",
                f"D) {self.question_table.item(row, 4).text()}",
                f"E) {self.question_table.item(row, 5).text()}"
            ]
            correct = self.question_table.item(row, 6).text()
            
            html += f"""
            <tr>
                <td>{row + 1}</td>
                <td>{question}</td>
                <td>{'<br>'.join(options)}</td>
                <td>{correct}</td>
            </tr>
            """
        
        html += """
        </table>
        <div class="footer">
            <p>Bu sınav Öğrenci Sınav Sistemi v2.0 ile oluşturulmuştur.</p>
        </div>
        </body>
        </html>
        """
        
        # PDF oluştur
        try:
            doc = QTextDocument()
            doc.setHtml(html)
            
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            
            doc.print_(printer)
            QMessageBox.information(self, "Başarılı", "PDF başarıyla oluşturuldu!")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulurken hata oluştu:\n{str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExamSystemApp()
    window.show()
    sys.exit(app.exec_())